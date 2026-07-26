#!/usr/bin/env python3
"""
generar_datos.py
----------------
Convierte el Excel de notas (hojas "NOTAS" y "Datos") en el archivo
data/estudiantes.json que usa el Portal de Notas.

Uso:
    python3 generar_datos.py "Notas_OFIAI.xlsx"

Cada vez que actualices las notas en el Excel, vuelve a ejecutar este
script y sube (git add / commit / push) el nuevo data/estudiantes.json
al repositorio de GitHub. La página se actualiza sola.

Requisitos:
    pip install openpyxl
"""

import sys
import json
import hashlib
from pathlib import Path

import openpyxl

MESES = {
    "ENE": "Enero", "FEB": "Febrero", "MAR": "Marzo", "ABR": "Abril",
    "MAY": "Mayo", "JUN": "Junio", "JUL": "Julio", "AGO": "Agosto",
    "SEP": "Septiembre", "SET": "Septiembre", "OCT": "Octubre",
    "NOV": "Noviembre", "DIC": "Diciembre",
}


def sha256_hex(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def normaliza_dni(valor) -> str:
    """DNI peruano = 8 dígitos. Si el Excel lo guardó como número y se
    comió un cero a la izquierda, lo restituye."""
    if valor is None:
        return ""
    return str(valor).strip().zfill(8)


def periodo_legible(mes, anio):
    if not mes:
        return ""
    mes_txt = MESES.get(str(mes).strip().upper(), str(mes).strip().title())
    if anio is None:
        return mes_txt
    anio_num = int(anio)
    anio_completo = 2000 + anio_num if anio_num < 100 else anio_num
    return f"{mes_txt} {anio_completo}"


def main():
    if len(sys.argv) < 2:
        print("Uso: python3 generar_datos.py <archivo_excel.xlsx>")
        sys.exit(1)

    ruta_excel = Path(sys.argv[1])
    if not ruta_excel.exists():
        print(f"No se encontró el archivo: {ruta_excel}")
        sys.exit(1)

    wb = openpyxl.load_workbook(ruta_excel, data_only=True)

    for hoja in ("NOTAS", "Datos"):
        if hoja not in wb.sheetnames:
            print(f"El Excel no tiene una hoja llamada '{hoja}'.")
            sys.exit(1)

    ws_notas = wb["NOTAS"]
    ws_datos = wb["Datos"]

    # --- indexar hoja Datos por DNI (para sacar correo, nombres, mes, año...) ---
    datos_por_dni = {}
    encabezados_datos = {ws_datos.cell(row=1, column=c).value: c
                          for c in range(1, ws_datos.max_column + 1)
                          if ws_datos.cell(row=1, column=c).value}

    col_dni_datos = encabezados_datos.get("DNI")
    for r in range(2, ws_datos.max_row + 1):
        dni_val = ws_datos.cell(row=r, column=col_dni_datos).value
        if dni_val is None:
            continue
        fila = {nombre: ws_datos.cell(row=r, column=c).value
                for nombre, c in encabezados_datos.items()}
        datos_por_dni[normaliza_dni(dni_val)] = fila

    # --- recorrer hoja NOTAS y armar cada estudiante ---
    encabezados_notas = {ws_notas.cell(row=2, column=c).value: c
                          for c in range(1, ws_notas.max_column + 1)
                          if ws_notas.cell(row=2, column=c).value}

    col_dni_notas = encabezados_notas.get("DNI")

    def val(fila, nombre_col):
        c = encabezados_notas.get(nombre_col)
        if c is None:
            return None
        v = ws_notas.cell(row=fila, column=c).value
        return v

    estudiantes = []
    dnis_vistos_en_notas = set()

    for r in range(3, ws_notas.max_row + 1):
        dni_val = val(r, "DNI")
        nombre_completo = val(r, "APELLIDOS Y NOMBRES")
        if dni_val is None and not nombre_completo:
            continue

        dni_norm = normaliza_dni(dni_val) if dni_val is not None else None
        info_extra = datos_por_dni.get(dni_norm, {}) if dni_norm else {}

        email = (info_extra.get("E-MAIL") or "").strip()
        if not email:
            # Sin correo no se puede armar el login de este alumno
            continue

        registro = {
            "email": email.lower(),
            "dni_hash": sha256_hex(dni_norm) if dni_norm else None,
            "nombre_completo": nombre_completo or info_extra.get("APELLIDOS Y NOMBRES"),
            "nombres": info_extra.get("NOMBRES"),
            "apellidos": info_extra.get("APELLIDOS"),
            "curso": val(r, "CURSO") or info_extra.get("CURSO"),
            "grupo": val(r, "GRUPO") or info_extra.get("GRUPO"),
            "periodo": periodo_legible(info_extra.get("MES"), info_extra.get("AÑO")),
            "asistencia": val(r, "ASISTE"),
            "asistencia_max": 6,
            "trabajo1": val(r, "TRABAJO 1"),
            "trabajo1_max": 30,
            "trabajo2": val(r, "TRABAJO 2"),
            "trabajo2_max": 30,
            "examen1": val(r, "EXAMEN 1"),
            "examen1_max": 40,
            "examen2": val(r, "EXAMEN 2"),
            "examen2_max": 40,
            "promedio_final": val(r, "PROM FIN"),
            "promedio_final_max": 20,
            "condicion": val(r, "CONDICION") or "RETIRADO",
        }
        estudiantes.append(registro)
        if dni_norm:
            dnis_vistos_en_notas.add(dni_norm)

    # --- alumnos que están en Datos pero ya no aparecen en NOTAS (retirados) ---
    for dni_norm, info_extra in datos_por_dni.items():
        if dni_norm in dnis_vistos_en_notas:
            continue
        email = (info_extra.get("E-MAIL") or "").strip()
        if not email:
            continue
        estudiantes.append({
            "email": email.lower(),
            "dni_hash": sha256_hex(dni_norm),
            "nombre_completo": info_extra.get("APELLIDOS Y NOMBRES"),
            "nombres": info_extra.get("NOMBRES"),
            "apellidos": info_extra.get("APELLIDOS"),
            "curso": info_extra.get("CURSO"),
            "grupo": info_extra.get("GRUPO"),
            "periodo": periodo_legible(info_extra.get("MES"), info_extra.get("AÑO")),
            "asistencia": None, "asistencia_max": 6,
            "trabajo1": None, "trabajo1_max": 30,
            "trabajo2": None, "trabajo2_max": 30,
            "examen1": None, "examen1_max": 40,
            "examen2": None, "examen2_max": 40,
            "promedio_final": None, "promedio_final_max": 20,
            "condicion": "RETIRADO",
        })

    salida = Path(__file__).parent / "data" / "estudiantes.json"
    salida.parent.mkdir(parents=True, exist_ok=True)
    salida.write_text(json.dumps(estudiantes, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Listo. {len(estudiantes)} alumnos escritos en {salida}")


if __name__ == "__main__":
    main()
